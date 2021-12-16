import time
import re
import pyautogui
import requests
import sys
import random
import pandas as pd
import imagehash
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime
from PIL import Image
from openpyxl.utils.dataframe import dataframe_to_rows

class MyLikes:
    def __init__(self, url, driver_path, records_path) -> None:
        self.url = url                      # URL for selenium
        self.incrementer = 0                # Variable to replace a count within a for loop for `main`
        self.card_identifier = dict()       # Unique identifier for a profile card
        self.picture_count = 0              # This helps to identify the profile card we're on and is also used in the filenames
        self.records = list()               # Storing the data to be written to an Excel workbook
        self.records_path = records_path    # Path to save the Excel workbook
        self.seen_cards = list()            # Which cards have already been seen by the script?
        self.now = datetime.utcnow()        # Store the start time, in GMT, of the script in a variable
        self.url_regEx = re.compile(pattern=r'url\(\"(https://.+\.jpg)\"')
        self.card_identifier_regEx = re.compile(pattern=r'https://images-ssl.gotinder.com/(.+)/\d{3}x')
        self.distance_regEx = re.compile(pattern=r'(\d+) miles away')
        self.options = webdriver.ChromeOptions() # Standard for using Chrome with selenium
        self.options.add_experimental_option('debuggerAddress', 'localhost:9222') # Running Chrome on localhost
        self.driver = webdriver.Chrome(executable_path=driver_path, options=self.options) # Standard for using Chrome with selenium
        self.headers = { # Headers for our requests. These are very important. Without them, we can get timed out or banned.
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36',
                    'accept-language': 'en-US,en;q=0.9',
                    'referer': 'https://tinder.com/',
                    'dnt': '1'
        }

    def __repr__(self) -> str:
        return 'This script is intended to download all of the pictures and videos from the profiles on the "Likes Sent" section of Tinder.'

    def load_workbook(self, initials=False) -> None:
        # Load the existing Excel workbook that's being used as a database
        self.workbook = openpyxl.load_workbook(filename=self.records_path)
        self.sheet = self.workbook['Datetime']
        
        if initials:
            # Save the current database to a variable
            self.existing_df = pd.read_excel(self.records_path, sheet_name='Datetime', header=0)
            self.picture_count_from_workbook = len(set(self.existing_df['Card_ID']))

    def close_workbook(self) -> None:
        # Save and close the Excel workbook
        self.workbook.save(filename=self.records_path)
        self.workbook.close()

    def log_in(self) -> None:
        # Open the URL in Chrome
        self.driver.get(url=self.url)
        time.sleep(4)
                                               
        # Click the Likes Sent button           
        self.driver.find_element_by_xpath(xpath='//a[@href="/app/my-likes"]').click() # Selecting by the href of an anchor (i.e., 'a') tag
        time.sleep(3)

    def main(self) -> None:
        # while 1:
        for _ in range(7):
            # Sleep
            time.sleep(3)

            # Get the current page's HTML
            final_html = self.driver.page_source

            # Create a soup object
            self.soup = BeautifulSoup(final_html, 'html.parser')

            # Find all profile cards within the current HTML
            cards = self.soup.find_all('div', {'aria-label': re.compile(pattern=r'.*'), 'class': 'Bdrs(8px) Bgz(cv) Bgp(c) StretchedBox'})

            # Find the div's id for the div that holds the profile cards. This is important because Tinder frequently changes this id, the class name, etc.
            div_id = self.soup.find('div', {'class': 'Sb(s) D(f) Jc(c) Fxd(c) Animtf(l) Animfm(f) Animdur(.75s) NetHeight(100%,--side-nav-bar-height)--ml H(100%) Ovy(s) Ovsb(n) Ovs(touch)'})['id']
            
            # Iterate over the profile cards
            for card in cards:  
                card_identifier = re.search(pattern=self.card_identifier_regEx, string=str(card)).group(1)

                if sum([1 for item in set(self.existing_df['Card_ID']) if item == card_identifier]) == 1 and card_identifier not in self.seen_cards:
                    # Add to seen cards list
                    self.seen_cards.append(card_identifier)

                    # Add the card ID to the dictionary
                    self.card_identifier.setdefault(card_identifier, 0)

                    # Increment the picture count
                    self.picture_count += 1

                    continue # Since the profile card ID is in the "Card_ID" column of the current database, skip this card and go to the next card
                elif self.card_identifier.get(card_identifier) is not None:
                    continue # Since the profile card ID is in the dictionary, skip this card and go to the next card
                else:        # Since we haven't gathered the profile data (i.e., the profile card data aren't in the current database and aren't in the current "records" list), gather now
                    # Click in the background
                    pyautogui.moveTo(x=1850, y=350, duration=0.1)
                    pyautogui.click()

                    # Add to seen cards list
                    self.seen_cards.append(card_identifier)

                    # Add the card ID to the dictionary
                    self.card_identifier.setdefault(card_identifier, 0)

                    # Increment the picture count
                    self.picture_count += 1

                    # Increment the picture count that originates from the Excel workbook only when we gather new data
                    self.picture_count_from_workbook += 1

                    # Click the relevant profile card           
                    if self.driver.find_element_by_xpath(xpath=f'//*[@id="{div_id}"]/div[2]/div[{self.picture_count}]/div/div/span/div') is not None: # Tinder may change the div the xpath relates to. I can probably write a regular expression to account for this, but I manually updated this one.
                        try:
                            self.driver.find_element_by_xpath(xpath=f'//*[@id="{div_id}"]/div[2]/div[{self.picture_count}]/div/div/span/div').click()
                        except Exception as e:
                            self.driver.find_element_by_xpath(xpath=f'//*[@id="{div_id}"]/div[2]/div[{self.picture_count}]/div/div/span/div[2]/video').click()
                    elif self.driver.find_element_by_xpath(xpath=f'//*[@id="{div_id}"]/div[2]/div[{self.picture_count}]/div/div') is not None:
                        self.driver.find_element_by_xpath(xpath=f'//*[@id="{div_id}"]/div[2]/div[{self.picture_count}]/div/div').click()
                    else:
                        # Finish the script by writing the data to a dataframe then appending data to an Excel workbook's worksheet. Finally, call `sys.exit()`
                        sys.exit('The script is complete. There are no more profile cards to go through.')

                    # Sleep 
                    time.sleep(1)

                    # Get HTML of the profile card
                    profile_html = self.driver.page_source

                    second_soup = BeautifulSoup(profile_html, 'html.parser')

                    # Try to get the name from the profile card
                    if second_soup.find('h1', {'class': 'Fz($xl) Fw($bold) Fxs(1) Fxw(w) Pend(8px) M(0) D(i)'}) is not None:
                        name = second_soup.find('h1', {'class': 'Fz($xl) Fw($bold) Fxs(1) Fxw(w) Pend(8px) M(0) D(i)'}).text.title()
                    else:
                        name = 'Name Not Found'

                    # This may be empty, but the span tag should always be there
                    age = second_soup.find('span', {'class': 'Whs(nw) Fz($l)'}).text

                    # Try to get the distance from me from the profile card
                    if second_soup.find('div', {'class': 'Fz($ms)'}) is not None:
                        row_text = str(second_soup.find('div', {'class': 'Fz($ms)'}))

                        if re.search(pattern=self.distance_regEx, string=row_text):
                            distance = re.search(pattern=self.distance_regEx, string=row_text).group(1)
                        else:
                            distance = 'Distance Not Found'
                    else:
                        distance = 'Distance Not Found'

                    # Try to get the bio from the profile card
                    if second_soup.find('div', {'class': 'P(16px) Us(t) C($c-secondary) BreakWord Whs(pl) Fz($ms)'}) is not None:
                        bio = second_soup.find('div', {'class': 'P(16px) Us(t) C($c-secondary) BreakWord Whs(pl) Fz($ms)'}).text
                    else:
                        bio = 'Bio Not Found'

                    # Try to get the passions from the profile card
                    if second_soup.find_all('div', {'class': 'Bdrs(100px) Bd D(ib) Va(m) Fz($xs) Mend(8px) Mb(8px) Px(8px) Py(4px) Bdc($c-secondary) C($c-secondary)'}) is not None:
                        passions = second_soup.find_all('div', {'class': 'Bdrs(100px) Bd D(ib) Va(m) Fz($xs) Mend(8px) Mb(8px) Px(8px) Py(4px) Bdc($c-secondary) C($c-secondary)'})

                        passions_text = ''

                        for passion in passions:
                            passions_text += passion.text + ','

                        passions_text = passions_text.strip(',')
                    else:
                        passions_text = 'Passions Not Found'

                    if passions_text == '':
                        passions_text = 'Passions Not Found'

                    # Try to get the "My Anthem" from the profile card
                    if second_soup.find('div', {'class': 'Mb(4px) Ell Fz($ms)'}) is not None:
                        song_title = second_soup.find('div', {'class': 'Mb(4px) Ell Fz($ms)'}).text
                        song_artist = second_soup.find('span', {'class': 'Mstart(4px) Ell'}).text
                    else:
                        song_title = 'No Song Title Found'
                        song_artist = 'No Song Artist Found'
              
                    # Get the total number of pages in the profile card
                    try:
                        number_of_pages = int(second_soup.find('button', {'class': 'bullet D(ib) Va(m) Cnt($blank)::a D(b)::a Cur(p) bullet--active H(4px)::a W(100%)::a Py(4px) Px(2px) W(100%) Bdrs(100px)::a Bgc(#fff)::a focus-background-style'}).text.split('/')[1])
                    except Exception:  
                        number_of_pages = 1 # If there's only one page, there won't be a button.

                    # Iterate over the number of pages
                    for i in range(0, number_of_pages, 1):
                        time.sleep(1)

                        page_html = self.driver.page_source

                        page_soup = BeautifulSoup(page_html, 'html.parser')

                        current_card = page_soup.find('span', {'class': 'keen-slider__slide Wc($transform) Fxg(1)', 'aria-hidden': 'false', 'style': re.compile(pattern=r'.+')})

                        vid = current_card.find('video', {'class': 'W(100%)'})

                        # Find appropriate URL
                        try:
                            if vid: 
                                vid = vid['src']
                                download_url = vid
                            else:
                                download_url = re.search(pattern=self.url_regEx, string=str(current_card)).group(1)
                        except Exception:
                            print(f'Could not find a download URL, {self.picture_count_from_workbook}, page {number_of_pages}')
                            continue # Couldn't find a download URL

                        # Send GET request
                        r = requests.get(url=download_url, headers=self.headers)     

                        # Content Type (i.e., image or video) and Last-Modified
                        content_type, res_last_mod = r.headers['Content-Type'], r.headers['Last-Modified']
                        res_last_mod = self.to_datetime_obj(date_str=res_last_mod)
                        time_diff = ':'.join(str(self.now - res_last_mod).split(':')[:2])
                        
                        # Write picture/video to disk
                        with open(file=f'./tinder_pics/{self.picture_count_from_workbook}_{name}_{i+1}.{download_url[-3:]}', mode='wb') as file:
                            file.write(r.content)   

                        # If the content is an image, create a hash
                        if download_url[-3:] == 'jpg':
                            hash = imagehash.average_hash(image=Image.open(fp=f'./tinder_pics/{self.picture_count_from_workbook}_{name}_{i+1}.{download_url[-3:]}'))

                        # Append data to list
                        self.records.append(
                                           (name, 
                                            age, 
                                            distance,
                                            bio,
                                            passions_text,
                                            song_title,
                                            song_artist,
                                            card_identifier, 
                                            content_type, 
                                            res_last_mod, 
                                            self.now, 
                                            time_diff, 
                                            str(hash))
                        ) # Convert hash from imagehash.ImageHash to string

                        # Resetting hash. This can be handled in a better way.
                        hash = '' 

                        # Check if we need to click to go to the next page
                        if i != (number_of_pages - 1):
                            pyautogui.moveTo(x=1250, y=400, duration=0.1)
                            pyautogui.click()
                            time.sleep(1)
                        else:
                            continue

                    # Click off the profile card         
                    pyautogui.moveTo(x=1850, y=350, duration=0.1)
                    pyautogui.click()
                    time.sleep(1)

            # Move down the webpage
            if self.incrementer == 0:
                pyautogui.moveTo(x=1850, y=350, duration=0.5)
                time.sleep(1)
                print(f'Run number: {self.incrementer} | {pyautogui.position()}')
                pyautogui.scroll(clicks=-2000)
                time.sleep(2.5)
                pyautogui.scroll(clicks=-280)
                time.sleep(1)
                self.incrementer += 1
            else:
                print(f'Run number: {self.incrementer} | {pyautogui.position()}')
                time.sleep(random.randint(2, 3))
                pyautogui.scroll(clicks=-755)
                time.sleep(random.randint(2, 3))
                self.incrementer += 1

    def to_datetime_obj(self, date_str) -> datetime.strptime:
        return datetime.strptime(date_str, '%a, %d %b %Y %H:%M:%S %Z') 
    
    def append_to_workbook(self) -> None:
        # Create dataframe from "self.records" list
        self.df_append = pd.DataFrame(data=self.records, 
                                      columns=['Name', 
                                               'Age', 
                                               'Distance',
                                               'Bio',
                                               'Passions',
                                               'Song_Title',
                                               'Song_Artist',
                                               'Card_ID',
                                               'Type',
                                               'Last_Mod_Date', 
                                               'Current_Date', 
                                               'Time_Diff', 
                                               'Hash']
)

        # Append each row from the dataframe to the Excel workbook's worksheet
        for row in dataframe_to_rows(df=self.df_append, index=False, header=False):
            self.sheet.append(row)
              
# Instantiate the class and call the necessary functions, inputting your arguments
if __name__ == '__main__':
    c = MyLikes(url='https://tinder.com/app/recs', 
                driver_path=r'C:\Users\Alex\Desktop\Python drivers\chromedriver.exe',
                records_path=r'C:\Users\Alex\Desktop\Date_Test.xlsx')
    c.load_workbook(initials=True)
    c.close_workbook()
    c.log_in()
    c.main()
    c.load_workbook(initials=False)
    c.append_to_workbook()
    c.close_workbook()
